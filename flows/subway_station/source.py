"""국가철도공단 전국 도시철도 역사 정보를 레일포털에서 받는다.

로그인도 서비스키도 필요 없다. 상세 페이지의 다운로드 버튼이 가리키는 주소를 그대로
부르면 xlsx 가 떨어진다.

**공공데이터포털 사본(`15093755`)을 쓰지 않는다.** 그쪽은 `atachFileYn` 이 `N` 이고
`atchFileId` 가 비어 있는 링크형이라, 법정동(`15063424`)에 쓴 2단계 첨부파일
다운로드가 404 로 끝난다. 포털이 파일을 들고 있지 않고 여기를 가리킬 뿐이다.
게다가 포털 사본은 `_20241231` 인데 여기 원본은 `_20260630` 으로 더 최신이다.

원문 한 행은 역이 아니라 **역 × 노선**이다. 강남역이 2호선과 신분당선 두 행으로
들어 있고, 그 둘의 좌표도 조금 다르다. 사용자가 `강남역 신분당선` 을 골라 그 근처
부스를 찾는 것이 목적이므로 접지 않고 그대로 담는다.

**원문이 주는 열을 다 담지 않는다.** 쓰임이 역명 검색과 좌표 두 가지뿐이라 영문명,
주소, 운영기관, 전화번호, 환승역구분, 데이터기준일자는 읽지 않는다. 지점 수집에서
픽닷이 지번 주소를 공짜로 받으면서도 담지 않는 것과 같은 기준이다. 쓰지 않는 값을
담으면 그 값이 틀렸을 때 누가 책임지는지가 흐려지고, 원문이 바뀔 때 따라 고칠 것만
늘어난다.

이름 교정은 여기서 하지 않는다. 여기는 사이트가 준 것만 담고 `normalize` 가 고친다.
값이 틀렸을 때 누구 잘못인가로 가르면 여기까지는 사이트 잘못이다.
"""

import io
import re
from dataclasses import dataclass

import httpx
import openpyxl
from prefect import get_run_logger, task

DETAIL_URL = "https://data.kric.go.kr/rips/M_01_01/detail.do?id=32"
DOWNLOAD_URL = "https://data.kric.go.kr/rips/dataset/download.file"

DATASET_ID = "32"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36"
)

# 실패해도 200 에 HTML 이 올 수 있다. xlsx 는 zip 이므로 앞 두 바이트로 가른다.
ZIP_MAGIC = b"PK"

# 헤더 첫 열 이름. 시트 구조가 바뀌면 여기서 걸린다.
HEADER = "역번호"

# 원문 열 이름 -> 우리 필드. 열 순서에 기대지 않는다. 순서로 읽으면 가운데에 열이
# 하나 끼었을 때 값이 통째로 한 칸씩 밀린 채로 조용히 적재된다.
FIELDS = {
    "역번호": "station_no",
    "역사명": "name",
    "노선번호": "line_no",
    "노선명": "line_name",
    "역위도": "lat",
    "역경도": "lon",
}


@dataclass(frozen=True)
class SourceStation:
    """원문 한 행. 역이 아니라 (역, 노선) 하나다.

    이름에 `역` 이 붙은 것과 안 붙은 것이 섞여 오고 노선명 표기도 무규칙이지만
    손대지 않는다. 고치는 것은 `normalize` 의 일이다.
    """

    station_no: str
    name: str
    line_no: str
    line_name: str
    lat: float
    lon: float


def _text(value: object) -> str:
    """셀 하나를 문자열로.

    **역번호가 str 652 / int 447 로 섞여 온다.** 숫자로 읽으면 `'0736'` 의 앞 0 이
    날아가 같은 노선의 다른 역과 부딪히므로 문자열로 통일한다. 정수로 온 값은
    파일 안에서 이미 앞 0 이 없는 값이라 `str()` 로 충분하다.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@task(retries=3, retry_delay_seconds=[10, 30, 60])
def fetch_rows() -> tuple[bytes, str]:
    """xlsx 본문과 파일 이름을 받는다.

    이름(예: `전체_도시철도역사정보_20260630.xlsx`)에 기준일자가 들어 있다. 적재할
    테이블 코멘트에 남겨 어느 스냅샷으로 만든 테이블인지 알 수 있게 한다.
    """
    logger = get_run_logger()

    with httpx.Client(
        timeout=180.0,
        follow_redirects=True,
        headers={"User-Agent": USER_AGENT},
    ) as client:
        body = client.get(
            DOWNLOAD_URL,
            params={"type": "filedata", "id": DATASET_ID, "operation": "1"},
            headers={"Referer": DETAIL_URL},
        )

    body.raise_for_status()

    if not body.content.startswith(ZIP_MAGIC):
        raise ValueError(
            "xlsx 가 아닌 응답을 받았습니다. 다운로드 경로가 바뀌었을 수 "
            f"있습니다: {body.content[:200]!r}"
        )

    dataset = _filename(body.headers.get("content-disposition", ""))
    logger.info("스냅샷 %s (%d bytes)", dataset or "(이름 없음)", len(body.content))

    return body.content, dataset


def _filename(disposition: str) -> str:
    """`Content-Disposition` 에서 파일 이름을 꺼낸다.

    **HTTP 헤더는 latin-1 로 디코딩되는데 이 서버는 utf-8 바이트를 그대로 실어
    보낸다.** 그대로 두면 한글이 깨진 채로 테이블 코멘트에 박히므로 되돌린다.
    """
    match = re.search(r'filename="?([^";]+)"?', disposition)
    if not match:
        return ""

    name = match.group(1).strip()
    try:
        return name.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


@task
def parse_rows(body: bytes) -> list[SourceStation]:
    """xlsx 를 행 목록으로. 이름이나 좌표가 없는 줄은 버린다."""
    logger = get_run_logger()

    workbook = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    header = [_text(cell) for cell in next(rows, ())]
    if HEADER not in header:
        raise ValueError(
            f"헤더에 {HEADER!r} 가 없습니다. 시트 구조가 바뀌었을 수 있습니다: "
            f"{header[:8]}"
        )

    missing = [name for name in FIELDS if name not in header]
    if missing:
        raise ValueError(f"원문에 없는 열이 있습니다: {missing}")

    index = {name: header.index(name) for name in FIELDS}

    stations: list[SourceStation] = []
    skipped = 0

    for row in rows:
        cell = lambda name: row[index[name]] if index[name] < len(row) else None  # noqa: E731

        name = _text(cell("역사명"))
        lat, lon = _text(cell("역위도")), _text(cell("역경도"))

        # 좌표가 이 데이터의 쓸모다. 1km 반경 매핑의 기준이 되므로 없는 행은
        # 담아도 쓰이지 않는다. 전량에서 결측이 0건이라 여기 걸리면 원문이 바뀐 것이다.
        if not name or not lat or not lon:
            skipped += 1
            continue

        try:
            latitude, longitude = float(lat), float(lon)
        except ValueError:
            skipped += 1
            continue

        stations.append(
            SourceStation(
                station_no=_text(cell("역번호")),
                name=name,
                line_no=_text(cell("노선번호")),
                line_name=_text(cell("노선명")),
                lat=latitude,
                lon=longitude,
            )
        )

    workbook.close()

    if skipped:
        logger.warning("이름이나 좌표가 없어 건너뛴 줄 %d개", skipped)

    logger.info("원문 %d행", len(stations))
    return stations
